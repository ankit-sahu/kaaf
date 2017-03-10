import openpyxl
from config.config import Config
from itertools import islice 

from openpyxl import load_workbook

class DataRef(object):
	
	# def __init__(self, file_ext):
	# 	self.file_ext = file_ext

	def __init__(self):
		self.wb = None
		self.sheet = None

	@classmethod	
	def get_sheet(self, sheet_name):
		self.wb = load_workbook(filename = Config.data + Config.data_ref_file)
		self.sheet = self.wb.get_sheet_by_name(sheet_name)
		return self.sheet

	@classmethod
	def read(self, sheet_name):

		cur_sheet = DataRef.get_sheet(sheet_name)

		total_row = cur_sheet.max_row + 1
		total_col = cur_sheet.max_column + 1

		list_of_entry = []
		for i in range(2, total_row):
			item = []
			for j in range(2, total_col):
				item.append(cur_sheet.cell(row=i,column=j).value)
			list_of_entry.append(item)
		return list_of_entry

	@classmethod
	def read_keys(self,sheet_name):
		
		cur_sheet = DataRef.get_sheet(sheet_name)

		total_row = cur_sheet.max_row + 1
		total_col = cur_sheet.max_column + 1

		list_of_keys = []
		for i in range(1,2):
			keys = []
			for j in range(2,total_col):
				keys.append(cur_sheet.cell(row = i, column  = j).value)
			# list_of_keys.append(key)
		return keys

